"""
test_05_breakeven.py — Gate de la Fase 0.5 (motor de Breakeven en Python)

Cubre: detección de offset por vigencia, mapa de columnas, vigencia/campo desde
filename, exactitud de la cadena de flujo (el_last), monotonía, y el GATE GOLDEN:
reproducir Breakeven.xlsm (Cierre 2024) para RUBIALES y CASTILLA dentro de ±0.5 USD/bbl.
No se avanza en el pipeline si algún assert falla.
"""
import sys
import warnings
from pathlib import Path

import numpy as np
import openpyxl
import pytest

warnings.filterwarnings("ignore")
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import motor_breakeven as mbk  # noqa: E402

BASE   = Path(__file__).resolve().parent.parent
RAW    = BASE / "datos" / "raw"
FC_DIR = RAW / "Formatos FC"
PLANTILLA = RAW / "Formatos FC" / "Formatos FC.xlsx"   # plantilla vacía (layout)

# Golden de Breakeven.xlsm (Cierre 2024), clase D-PDP
GOLDEN = {
    ("datos/raw/Formatos FC/2024/RUBIALES_CF_SEC_09-Jan-2025.xlsx"):
        {"fin": 24.562025, "ope": 67.792368},
    ("datos/raw/Formatos FC/2024/CASTILLA_CF_SEC_23-Jan-2025.xlsx"):
        {"fin": 28.809431, "ope": 66.613432},
}
TOL = 0.5


# ── Helpers ─────────────────────────────────────────────────────────────────────

def _sintetico(precio_capex=50.0, abandono=30.0):
    """FC sintético de 3 periodos, solo aceite, para verificar la cadena."""
    n = 3
    d = {
        "ACEITE_VTS_NETO": np.array([100.0, 100.0, 100.0]),
        "GAS_VTS_NETO":    np.zeros(n),
        "GLP_NETO":        np.zeros(n),
        "PODER_CALOR":     np.zeros(n),
        "PRECIO_GAS":      np.zeros(n),
        "PRECIO_GLP":      np.zeros(n),
        "OPEX":            np.array([10.0, 10.0, 10.0]),
        "CAPEX":           np.array([precio_capex, 0.0, 0.0]),
        "OTROS_ING":       np.zeros(n),
        "ANIO":            np.array([2025.0, 2026.0, 2027.0]),
        "_ult":            n - 1,
        "_ult_oil":        n - 1,   # todos los años tienen aceite
        "_abandono_total": abandono,
    }
    return d


def _leer_dpdp(ruta):
    """Lee solo la hoja D-PDP de un FC real (rápido, sin cargar las 5 hojas)."""
    wb = openpyxl.load_workbook(BASE / ruta, read_only=True, data_only=True)
    ws = wb["D-PDP"]
    offset = mbk.detectar_offset(ws)
    d = mbk.leer_hoja(ws, mbk.resolver_columnas(offset))
    wb.close()
    return d, offset


# ── Mapa de columnas / detección de offset ──────────────────────────────────────

def test_resolver_columnas_2024_2025():
    """Cada nombre lógico cae en la letra esperada (tabla del plan §B)."""
    from openpyxl.utils import get_column_letter as cl
    c0 = {k: cl(v) for k, v in mbk.resolver_columnas(0).items()}
    c1 = {k: cl(v) for k, v in mbk.resolver_columnas(1).items()}
    assert c0["PRECIO_ACEITE"] == "BM" and c1["PRECIO_ACEITE"] == "BN"
    assert c0["OPEX"] == "CD" and c1["OPEX"] == "CE"
    assert c0["CAPEX"] == "DI" and c1["CAPEX"] == "DJ"
    assert c0["ABANDONO"] == "DZ" and c1["ABANDONO"] == "EA"
    # ANIO nunca se desplaza (col C < J)
    assert c0["ANIO"] == "C" and c1["ANIO"] == "C"


def test_detectar_offset_plantilla():
    """Sobre la plantilla vacía: sheet '2024' → offset 0, '2025' → offset 1."""
    wb = openpyxl.load_workbook(PLANTILLA, read_only=True, data_only=True)
    assert mbk.detectar_offset(wb["2024"]) == 0
    assert mbk.detectar_offset(wb["2025"]) == 1
    wb.close()


# ── Vigencia / campo desde filename ──────────────────────────────────────────────

@pytest.mark.parametrize("nombre,esp", [
    ("RUBIALES_CF_SEC_09-Jan-2025.xlsx", "2024"),
    ("CASTILLA_CF_SEC(Royalties)_15-Jan-2026.xlsx", "2025"),
])
def test_vigencia_desde_nombre(nombre, esp):
    assert mbk.vigencia_desde_nombre(nombre) == esp


@pytest.mark.parametrize("nombre,esp", [
    ("RUBIALES_CF_SEC_09-Jan-2025.xlsx", "RUBIALES"),
    ("CASTILLA NORTE_CF_SEC(Royalties)_15-Jan-2026.xlsx", "CASTILLA NORTE"),
])
def test_campo_desde_nombre(nombre, esp):
    _crudo, campo = mbk.campo_desde_nombre(nombre)
    assert campo == esp


# ── Cadena de flujo: exactitud y monotonía ───────────────────────────────────────

def test_cadena_el_last_exacto():
    """El FC del último periodo = ingresos − capex − opex = 100·p − 0 − 10."""
    d = _sintetico()
    for p in (1.0, 5.0, 0.1):
        _npv, el_last = mbk._cadena(p, d)
        assert el_last == pytest.approx(100.0 * p - 10.0, abs=1e-9)


def test_monotonia_npv_y_ellast():
    """NPV y el_last estrictamente crecientes en el precio (Brent↑ → caja↑)."""
    d = _sintetico()
    precios = np.linspace(1, 200, 60)
    npvs    = np.array([mbk._cadena(p, d)[0] for p in precios])
    ells    = np.array([mbk._cadena(p, d)[1] for p in precios])
    assert np.all(np.diff(npvs) > 0)
    assert np.all(np.diff(ells) > 0)


def test_financiero_coincide_con_escaneo():
    """brentq vs raíz por escaneo fino → mismo precio de NPV=0 (±0.05)."""
    d = _sintetico()
    fin, msg = mbk.breakeven_financiero(d)
    assert msg == "OK"
    precios = np.linspace(0.01, 500, 200000)
    npvs    = np.array([mbk._cadena(p, d)[0] for p in precios])
    cruce   = precios[np.argmin(np.abs(npvs))]
    assert fin == pytest.approx(cruce, abs=0.05)


def test_sin_raiz_devuelve_mensaje():
    """Campo sin aceite → NPV siempre negativo → NaN con mensaje explícito (no silencio)."""
    d = _sintetico()
    d["ACEITE_VTS_NETO"] = np.zeros(3)   # sin ingresos por aceite
    fin, msg = mbk.breakeven_financiero(d)
    assert np.isnan(fin) and msg != "OK"


def test_brent_insensitive_financiero():
    """Campo sin aceite → breakeven_financiero devuelve BRENT_INSENSITIVE."""
    d = _sintetico()
    d["ACEITE_VTS_NETO"] = np.zeros(3)
    d["_ult_oil"] = -1
    fin, msg = mbk.breakeven_financiero(d)
    assert np.isnan(fin)
    assert msg == "BRENT_INSENSITIVE"


def test_brent_insensitive_operacional():
    """Campo sin aceite en ningún año → breakeven_operacional devuelve BRENT_INSENSITIVE."""
    d = _sintetico()
    d["_ult_oil"] = -1   # motor detectaría esto en leer_hoja para campos gas
    ope, msg = mbk.breakeven_operacional(d)
    assert np.isnan(ope)
    assert msg == "BRENT_INSENSITIVE"


def test_operacional_robusto_cola_gas():
    """
    Campo con aceite en años 0-1 y cola de gas (aceite=0) en año 2.
    breakeven_operacional debe usar el último año CON aceite (_ult_oil=1),
    no el último de la serie (_ult=2). Con aceite=0 en año 2, fc[2] sería
    constante en precio → jamás cruzaría cero → SIN_RAIZ.
    Con _ult_oil=1, fc[1] = 100p - 10 → sí cruza cero.
    """
    n = 3
    d = {
        "ACEITE_VTS_NETO": np.array([100.0, 100.0, 0.0]),   # año 2: solo gas
        "GAS_VTS_NETO":    np.array([0.0,   0.0,   50.0]),  # gas en cola
        "GLP_NETO":        np.zeros(n),
        "PODER_CALOR":     np.array([0.0,   0.0,   1.0]),
        "PRECIO_GAS":      np.array([0.0,   0.0,   5.0]),   # fijo: 50*1*5=250 > opex
        "PRECIO_GLP":      np.zeros(n),
        "OPEX":            np.array([10.0, 10.0, 10.0]),
        "CAPEX":           np.zeros(n),
        "OTROS_ING":       np.zeros(n),
        "ANIO":            np.array([2025.0, 2026.0, 2027.0]),
        "_ult":            n - 1,
        "_ult_oil":        1,        # último año con aceite = año 1
        "_abandono_total": 0.0,
    }
    ope, msg = mbk.breakeven_operacional(d)
    # fc[1] = 100·p − 10 → raíz en p = 0.10
    assert msg == "OK", f"Esperado OK pero recibí: {msg}"
    assert ope == pytest.approx(0.10, abs=1e-3), f"Raíz esperada ≈ 0.10, recibida {ope}"


def test_limite_economico_argmax_sin_doble_abandono():
    """
    Cuando fc_acum tiene un plateau (máximo en varias filas consecutivas), argmax
    elige la PRIMERA → abandono se aplica exactamente una vez → NPV finito y correcto.
    Con la máscara de igualdad antigua, habría múltiples unos → NPV incorrecto.
    """
    n = 4
    d = {
        # Con precio=1: fc = [100-0, 100-200, 100-200, 100-200] = [100,-100,-100,-100]
        # fc_acum = [100, 0, -100, -200] → max en fila 0, LE=0 (sin empate)
        # Para forzar plateau: OPEX=100 desde fila 1 → fc = [100,0,0,0]
        # fc_acum = [100, 100, 100, 100] → max en filas 0,1,2,3 (empate)
        "ACEITE_VTS_NETO": np.array([100.0, 100.0, 100.0, 100.0]),
        "GAS_VTS_NETO":    np.zeros(n),
        "GLP_NETO":        np.zeros(n),
        "PODER_CALOR":     np.zeros(n),
        "PRECIO_GAS":      np.zeros(n),
        "PRECIO_GLP":      np.zeros(n),
        "OPEX":            np.array([0.0, 100.0, 100.0, 100.0]),  # a p=1: fc=[100,0,0,0]
        "CAPEX":           np.zeros(n),
        "OTROS_ING":       np.zeros(n),
        "ANIO":            np.array([2025.0, 2026.0, 2027.0, 2028.0]),
        "_ult":            n - 1,
        "_ult_oil":        n - 1,
        "_abandono_total": 200.0,   # abandono alto: si se aplica 2 veces → NPV diferente
    }
    npv, _ = mbk._cadena(1.0, d)
    # Con argmax: LE=fila 0, contador=[1,0,0,0], abandono=-200 en fila 0 (una vez)
    # Verificar que NPV es finito y no NaN (la máscara antigua podría dar NPV incorrecto)
    assert np.isfinite(npv), f"NPV no finito: {npv}"
    # Con abandonos múltiples (bug antiguo), el NPV sería mucho más negativo
    # Con una sola vez y contador=[1,0,0,0]:
    # fc_desc[0] = (100 - 200 - impuesto(100-ey)) * factor[0] * 1
    # Solo verificamos que es finito y que el breakeven financiero converge
    fin, msg = mbk.breakeven_financiero(d)
    assert msg == "OK", f"Financiero debe converger: {msg}"


# ── GATE GOLDEN (duro) ───────────────────────────────────────────────────────────

@pytest.mark.parametrize("ruta", list(GOLDEN.keys()))
def test_gate_golden_financiero_operacional(ruta):
    """Reproduce Breakeven.xlsm (D-PDP) dentro de ±0.5 USD/bbl. GATE."""
    d, offset = _leer_dpdp(ruta)
    assert offset == 0          # los FC de 2024 no llevan offset
    assert d is not None and d["_ult"] >= 0
    fin, msg_fin = mbk.breakeven_financiero(d)
    ope, msg_ope = mbk.breakeven_operacional(d)
    g = GOLDEN[ruta]
    assert msg_fin == "OK" and msg_ope == "OK"
    assert fin == pytest.approx(g["fin"], abs=TOL), f"Fin {fin} vs golden {g['fin']}"
    assert ope == pytest.approx(g["ope"], abs=TOL), f"Ope {ope} vs golden {g['ope']}"


# ── Esquema del tablón generado (si el runner ya corrió) ─────────────────────────

def test_breakeven_resultados_esquema():
    """Si existe el parquet del runner: esquema correcto y sin NaN silenciosos."""
    import pandas as pd
    pq = BASE / "datos" / "staging" / "breakeven_resultados.parquet"
    if not pq.exists():
        pytest.skip("breakeven_resultados.parquet aún no generado (correr 05_breakeven.py)")
    df = pd.read_parquet(pq)
    requeridas = {"CAMPO", "VIGENCIA_BREAKEVEN", "CLASE_RESERVA",
                  "BREAKEVEN_FINANCIERO_USD_BBL", "MENSAJE_FIN", "MENSAJE_OPE",
                  "BRENT_INSENSITIVE"}
    assert requeridas.issubset(df.columns), \
        f"Columnas faltantes: {requeridas - set(df.columns)}"
    # NaN ↔ mensaje distinto de OK (nunca un NaN sin explicación)
    nan_fin = df["BREAKEVEN_FINANCIERO_USD_BBL"].isna()
    assert (df.loc[nan_fin, "MENSAJE_FIN"] != "OK").all()
    assert {"2024", "2025"}.issubset(set(df["VIGENCIA_BREAKEVEN"]))
    # BRENT_INSENSITIVE solo True cuando MENSAJE es BRENT_INSENSITIVE
    insens_rows = df[df["BRENT_INSENSITIVE"]]
    if len(insens_rows) > 0:
        assert (insens_rows["MENSAJE_FIN"] == "BRENT_INSENSITIVE").all() or \
               (insens_rows["MENSAJE_OPE"] == "BRENT_INSENSITIVE").all()

"""
motor_breakeven.py — Motor de Breakeven en Python (port del VBA `Codigo Breakeven`)

POR QUÉ: El Breakeven Financiero (precio NET donde NPV=0) y el Operacional (precio
donde el FC del último año = 0) son inputs del pipeline de predicción 1P. Antes se
generaban con una macro Excel (Solver + GoalSeek) que falla en ~80% de los campos,
exige clicks manuales y hardcodea letras de columna. Este motor reproduce la misma
matemática de forma headless, determinista (raíz por `brentq`) y testeable.

VALIDADO contra `Breakeven.xlsm` (golden Cierre 2024): RUBIALES/CASTILLA D-PDP
reproducen Financiero y Operacional dentro de ±0.01 USD/bbl.

CAMBIO DE FORMATO 2025: el FC Cierre-2025 insertó la columna "Pozos iny vapor" en la
posición J, desplazando +1 toda columna ≥ J (las A–I, incl. C=Año, no se mueven). El
motor detecta el offset por encabezado, así que sirve para ambas vigencias sin tocar
código.
"""

import re
from pathlib import Path

import numpy as np
import openpyxl
from openpyxl.utils import column_index_from_string as _ci
from openpyxl.utils import get_column_letter as _cl
from scipy.optimize import brentq

from homologacion import Homologador

# ── Constantes del modelo (escritas por el VBA PrepararArchivos_V24) ────────────
SENSIBILIDAD_OPEX = 0.0    # EO6  — sensibilidad opex (0 = sin sensibilizar)
TASA_DESCUENTO    = 0.10   # EQ9  — 10% anual
TASA_IMPUESTOS    = 0.35   # EU11 — 35%

PRIMERA_FILA = 12          # firstDataRow
ULTIMA_FILA  = 70          # lastDataRowLimit (búsqueda de última fila con datos)

# Hojas de clase de reserva a procesar (lista del VBA V24)
HOJAS_CLASE = ["D-PDP", "D-PNP", "D-PND", "D-PRB", "D-PS"]

# Bracket de precio NET (USD/bbl) para el root-find. Amplio para cubrir cualquier campo.
BRACKET = (0.01, 500.0)

# ── Mapa de columnas base (vigencia 2024, índices 1-based) ──────────────────────
# Nombre lógico → letra 2024. ANIO no se mueve (col C < J); el resto shiftea +offset.
COLS_2024 = {
    "ANIO":            "C",   # LEFT(C,4) = año del periodo
    "ACEITE_VTS_NETO": "BE",  # Aceite Ventas Neto ECP (barriles) — × EL3
    "GAS_VTS_NETO":    "BG",  # Gas Ventas Neto ECP (KPC)
    "GLP_NETO":        "BI",  # GLP Neto ECP (barriles)
    "PRECIO_ACEITE":   "BM",  # Precio Aceite (ancla de detección de offset)
    "PODER_CALOR":     "BP",  # Poder Calorífico (MMBTU/MCF)
    "PRECIO_GAS":      "BQ",  # Precio Gas (USD/MMBTU)
    "PRECIO_GLP":      "BS",  # Precio GLP (USD/barril)
    "OPEX":            "CD",  # OpEx Ecopetrol (USD)
    "CAPEX":           "DI",  # CAPEX ECP (dólares)
    "ABANDONO":        "DZ",  # Costo Abandono (dólares)
    "OTROS_ING":       "EG",  # Otros Ingresos-Egresos (dólares)
}
# Encabezados (fila 10) usados para detectar el offset por vigencia
_HDR_ANCLA = {
    "PRECIO_ACEITE": "PRECIO ACEITE",
    "OPEX":          "OPEX ECOPETROL",
    "CAPEX":         "CAPEX ECP",
}

_H = Homologador()


# ── Helpers ─────────────────────────────────────────────────────────────────────

def _norm(s) -> str:
    """Normaliza encabezado: mayúsculas, solo [A-Z0-9 ], colapsa espacios.
    Hace converger mojibake (Calor�fico) y acentos (Calorífico) al mismo token."""
    t = re.sub(r"[^A-Z0-9 ]", "", str(s).upper())
    return re.sub(r"\s+", " ", t).strip()


def detectar_offset(ws) -> int:
    """
    Detecta el desplazamiento de columnas de la hoja FC respecto al layout 2024.
    Ancla = 'Precio Aceite' (encabezado único en fila 10). Verifica con OpEx/CAPEX.
    Devuelve 0 (Cierre 2024) o 1 (Cierre 2025). Cualquier otro → error duro.
    """
    base_precio = _ci(COLS_2024["PRECIO_ACEITE"])
    encontrada = None
    # Buscar 'PRECIO ACEITE' exacto en fila 10 cerca del rango esperado
    for col in range(base_precio - 3, base_precio + 5):
        if _norm(ws.cell(10, col).value) == _HDR_ANCLA["PRECIO_ACEITE"]:
            encontrada = col
            break
    if encontrada is None:
        raise ValueError("No se encontró el encabezado 'Precio Aceite' (fila 10).")
    offset = encontrada - base_precio
    if offset not in (0, 1):
        raise ValueError(f"Offset de columnas inesperado: {offset} (se esperaba 0 o 1).")
    # Verificación cruzada con dos anclas más
    for nombre in ("OPEX", "CAPEX"):
        col = _ci(COLS_2024[nombre]) + offset
        if _norm(ws.cell(10, col).value) != _HDR_ANCLA[nombre]:
            raise ValueError(
                f"Offset {offset} inconsistente: '{nombre}' no cae en {_cl(col)}.")
    return offset


def resolver_columnas(offset: int) -> dict:
    """Mapa nombre lógico → índice de columna 1-based para el offset dado.
    ANIO (col C) nunca se desplaza; el resto sí (todas ≥ columna J)."""
    out = {}
    for nombre, letra in COLS_2024.items():
        idx = _ci(letra)
        out[nombre] = idx if nombre == "ANIO" else idx + offset
    return out


def vigencia_desde_nombre(nombre_archivo: str) -> str:
    """`*_CF_SEC_*-<mes>-<AAAA>[_vN].xlsx` → VIGENCIA = AAAA − 1 (Jan-2025 = Cierre 2024).
    El sufijo opcional `_v2`, `_v3`, etc. se ignora antes de buscar el año."""
    # Eliminar sufijo de versión (_v2, _v3, ...) antes del .xlsx
    nombre_limpio = re.sub(r"_v\d+\.xlsx$", ".xlsx", nombre_archivo, flags=re.IGNORECASE)
    m = re.search(r"-(\d{4})\.xls", nombre_limpio)
    if not m:
        raise ValueError(f"No se pudo extraer el año de cierre de '{nombre_archivo}'.")
    return str(int(m.group(1)) - 1)


def campo_desde_nombre(nombre_archivo: str) -> tuple[str, str]:
    """Extrae el nombre crudo de campo (antes de _CF_SEC) y lo homologa."""
    crudo = re.split(r"_CF_SEC", nombre_archivo)[0].strip()
    campo, _flag = _H.homologar(crudo)
    return crudo, campo


# ── Lectura de una hoja FC ──────────────────────────────────────────────────────

def leer_hoja(ws, cols: dict) -> dict | None:
    """
    Carga las columnas fuente de una hoja FC (filas 12..70) en arrays numpy.
    Devuelve None si no hay categoría de reservas (col Precio Aceite toda en 0).
    `ult` = última fila (0-based dentro del bloque) con OpEx ≠ 0.

    Usa iter_rows para recorrer el XML una sola vez (evita O(n) re-scans de
    ws.cell() en modo read_only, que causaban ~90 s por hoja).
    """
    n = ULTIMA_FILA - PRIMERA_FILA + 1
    datos = {k: np.zeros(n) for k in cols if k != "ANIO"}
    anios = np.zeros(n)

    # Mapa inverso: índice de columna → nombre lógico (solo los que necesitamos)
    col_a_nombre = {c: k for k, c in cols.items()}
    min_col = min(cols.values())
    max_col = max(cols.values())

    for row_idx, row in enumerate(
            ws.iter_rows(min_row=PRIMERA_FILA, max_row=ULTIMA_FILA,
                         min_col=min_col, max_col=max_col)):
        for cell in row:
            nombre = col_a_nombre.get(cell.column)
            if nombre is None:
                continue
            v = cell.value
            if nombre == "ANIO":
                try:
                    anios[row_idx] = int(str(v)[:4])
                except (ValueError, TypeError):
                    anios[row_idx] = 0
            else:
                datos[nombre][row_idx] = v if isinstance(v, (int, float)) else 0.0

    datos["ANIO"] = anios

    if np.count_nonzero(datos["PRECIO_ACEITE"]) == 0:
        return None  # Sin categoría de reservas (gate foundValueInBM del VBA)

    nz = np.nonzero(datos["OPEX"])[0]
    datos["_ult"] = int(nz.max()) if nz.size else -1
    # Abandono total = SUM(DZ12:DZ70) sobre TODO el rango (puede caer tras el LE)
    datos["_abandono_total"] = float(datos["ABANDONO"].sum())

    # Último año con producción de aceite dentro del horizonte activo.
    # Usado por breakeven_operacional: en colas de gas, los últimos años tienen
    # aceite=0 y fc[-1] sería constante en precio → no cruzaría cero nunca.
    ult = datos["_ult"]
    nz_oil = (np.nonzero(datos["ACEITE_VTS_NETO"][:ult + 1])[0]
              if ult >= 0 else np.array([], dtype=int))
    datos["_ult_oil"] = int(nz_oil.max()) if nz_oil.size else -1
    return datos


# ── Combinación multi-FC (directriz §4bis.3, 2026-07-11) ────────────────────────

def combinar_hojas(ds: list[dict]) -> dict:
    """
    Combina los datos de N hojas FC (misma clase de reserva) de componentes físicos de
    un mismo campo UNIFICADO en UN solo flujo de caja, alineado por AÑO CALENDARIO.

    POR QUÉ: los componentes certifican juntos (doctrina Agregación v3) → son un solo
    negocio económico. Calcular el breakeven de cada componente y quedarse con uno
    (bug de primera fila) o promediarlos NO da el breakeven del campo combinado: la
    raíz de una suma de flujos no es el promedio de las raíces.

    Reglas de combinación:
      SUMA         : ACEITE_VTS_NETO, GAS_VTS_NETO, GLP_NETO, OPEX, CAPEX, ABANDONO,
                     OTROS_ING, _abandono_total.
      EQUIVALENTE  : el ingreso de gas se preserva EXACTO vía
                     PRECIO_GAS_comb = Σ(BGᵢ·BPᵢ·BQᵢ)/Σ(BGᵢ) con PODER_CALOR=1
                     (el producto de promedios ≠ promedio de productos). Igual GLP:
                     PRECIO_GLP_comb = Σ(BIᵢ·BSᵢ)/Σ(BIᵢ).
      RECALCULA    : _ult (última fila con OPEX≠0) y _ult_oil sobre el combinado.

    Los años activos (fila ≤ _ult) de cada componente deben tener AÑO válido (>0);
    de lo contrario la alineación calendario es imposible → error duro, nunca silencio.
    """
    if not ds:
        raise ValueError("combinar_hojas: lista vacía.")
    if len(ds) == 1:
        return ds[0]

    # Años activos por componente (validación: sin años inválidos dentro del horizonte)
    anios_union = set()
    for d in ds:
        ult = d["_ult"]
        if ult < 0:
            continue
        anios_comp = d["ANIO"][:ult + 1]
        if np.any(anios_comp <= 0):
            raise ValueError("combinar_hojas: fila activa sin AÑO válido — "
                             "alineación por año calendario imposible.")
        anios_union.update(int(a) for a in anios_comp)
    if not anios_union:
        raise ValueError("combinar_hojas: ningún componente con horizonte activo.")

    anios = np.array(sorted(anios_union), dtype=float)
    n = len(anios)
    idx_de_anio = {int(a): i for i, a in enumerate(anios)}

    SUMAR = ["ACEITE_VTS_NETO", "GAS_VTS_NETO", "GLP_NETO",
             "OPEX", "CAPEX", "ABANDONO", "OTROS_ING"]
    comb = {k: np.zeros(n) for k in SUMAR}
    ing_gas = np.zeros(n)   # Σ BGᵢ·BPᵢ·BQᵢ por año (para el precio equivalente)
    ing_glp = np.zeros(n)   # Σ BIᵢ·BSᵢ por año
    abandono_total = 0.0

    for d in ds:
        ult = d["_ult"]
        abandono_total += float(d.get("_abandono_total", 0.0))
        if ult < 0:
            continue
        for i in range(ult + 1):
            j = idx_de_anio[int(d["ANIO"][i])]
            for k in SUMAR:
                if k in d:   # ABANDONO por fila puede faltar (el total va en _abandono_total)
                    comb[k][j] += float(d[k][i])
            ing_gas[j] += float(d["GAS_VTS_NETO"][i] * d["PODER_CALOR"][i]
                                * d["PRECIO_GAS"][i])
            ing_glp[j] += float(d["GLP_NETO"][i] * d["PRECIO_GLP"][i])

    # Precios equivalentes que preservan el ingreso combinado exacto
    with np.errstate(divide="ignore", invalid="ignore"):
        precio_gas = np.where(comb["GAS_VTS_NETO"] > 0,
                              ing_gas / comb["GAS_VTS_NETO"], 0.0)
        precio_glp = np.where(comb["GLP_NETO"] > 0,
                              ing_glp / comb["GLP_NETO"], 0.0)

    comb["PODER_CALOR"] = np.where(comb["GAS_VTS_NETO"] > 0, 1.0, 0.0)
    comb["PRECIO_GAS"]  = precio_gas
    comb["PRECIO_GLP"]  = precio_glp
    # PRECIO_ACEITE solo se usa como gate de "hay reservas" en leer_hoja; aquí basta
    # un marcador no-cero en los años con aceite.
    comb["PRECIO_ACEITE"] = np.where(comb["ACEITE_VTS_NETO"] > 0, 1.0, 0.0)
    comb["ANIO"] = anios

    nz = np.nonzero(comb["OPEX"])[0]
    comb["_ult"] = int(nz.max()) if nz.size else -1
    ult = comb["_ult"]
    nz_oil = (np.nonzero(comb["ACEITE_VTS_NETO"][:ult + 1])[0]
              if ult >= 0 else np.array([], dtype=int))
    comb["_ult_oil"] = int(nz_oil.max()) if nz_oil.size else -1
    comb["_abandono_total"] = abandono_total
    return comb


# ── Núcleo: cadena de flujo de caja (reproduce las fórmulas EK..EV del VBA) ──────

def _cadena(precio: float, d: dict) -> tuple[float, float]:
    """
    Para un precio NET de aceite (EL3 flat), calcula:
      - NPV  = SUM(Flujo de Caja Descontado)  [EV11]
      - el_ultimo = Flujo de Caja del último periodo [EL en la última fila]
    El aceite usa el precio flat (EL3); gas/GLP usan sus precios por periodo.
    """
    s = slice(0, d["_ult"] + 1)
    BE, BG, BI = d["ACEITE_VTS_NETO"][s], d["GAS_VTS_NETO"][s], d["GLP_NETO"][s]
    BP, BQ, BS = d["PODER_CALOR"][s], d["PRECIO_GAS"][s], d["PRECIO_GLP"][s]
    CD, DI, EG = d["OPEX"][s], d["CAPEX"][s], d["OTROS_ING"][s]
    anio = d["ANIO"][s]

    opex     = (CD + EG) * (1 - SENSIBILIDAD_OPEX)
    ingresos = BE * precio + BG * BP * BQ + BI * BS        # EK (después de regalías)
    fc       = ingresos - DI - opex                        # EL = Flujo de Caja
    fc_acum  = np.cumsum(fc)                               # EM

    # Límite Económico: primera fila del máximo acumulado.
    # Se usa argmax en lugar de comparación de igualdad para evitar el bug de empate:
    # si dos filas tienen el mismo máximo, la máscara booleana antigua producía múltiples
    # unos → contador negativo → abandono restado varias veces → NPV incorrecto.
    le_idx   = int(np.argmax(fc_acum))                     # EN — primera fila del máx
    limite   = np.zeros_like(fc_acum)
    limite[le_idx] = 1.0                                   # exactamente un 1
    contador = 1 - np.cumsum(limite) + limite              # EP (1 hasta LE incl., 0 después)

    abandono = -d["_abandono_total"] * limite              # EQ (aplicado una sola vez en LE)
    factor   = 1.0 / (1 + TASA_DESCUENTO) ** (anio - anio[0] + 0.5)   # ER

    # Depreciación para impuestos (EW·EX = EY); columnas ES/ET del VBA son inertes
    ew11 = float(np.sum(ingresos * contador))
    ew   = np.cumsum(ingresos) / ew11 if ew11 != 0 else np.zeros_like(ingresos)
    ex   = np.cumsum(DI * contador)
    ey   = ew * ex

    base_imp  = ingresos - opex - ey
    impuestos = np.where(base_imp < 0, 0.0, base_imp * TASA_IMPUESTOS)   # EU
    fc_desc   = (fc + abandono - impuestos) * factor * contador          # EV

    # el_ope: FC del último año CON aceite (= precio de límite económico / abandono).
    # En colas de gas, los años finales tienen ACEITE=0 y fc[-1] sería constante en
    # precio → no cruzaría cero → el breakeven sería irresoluble. Usar el último año
    # con aceite como referencia garantiza convergencia mientras el campo tenga aceite.
    ult_oil = d.get("_ult_oil", len(fc) - 1)
    el_ope  = float(fc[ult_oil]) if ult_oil >= 0 else float(fc[-1])

    return float(fc_desc.sum()), el_ope


def _raiz(funcion, brent_insensible: bool = False) -> tuple[float, str]:
    """brentq sobre [BRACKET] si hay cambio de signo. Devuelve (precio, mensaje).

    brent_insensible=True: si no hay cruce, devuelve BRENT_INSENSITIVE en vez de
    SIN_RAIZ_*. Indica que el campo no depende del precio Brent (p.ej. dominado por
    ingresos de gas con precio fijo), no que el solver haya fallado.

    Raíz de borde (2026-07-11, directriz §4bis.4): una raíz pegada al borde inferior
    del bracket (≤ lo + 0.01) significa que el objetivo es ≥ 0 a cualquier precio
    positivo — el precio de equilibrio económico verdadero es 0, no un artefacto del
    bracket numérico. Se devuelve (0.0, "OK_SIEMPRE_RENTABLE") para que sea visible.
    """
    lo, hi = BRACKET
    try:
        flo, fhi = funcion(lo), funcion(hi)
    except (ValueError, ZeroDivisionError, FloatingPointError):
        return np.nan, "ERROR_CALCULO"
    if np.isnan(flo) or np.isnan(fhi):
        return np.nan, "ERROR_CALCULO"
    if flo * fhi > 0:
        # Sin cruce de signo en el bracket
        if brent_insensible:
            # El ingreso de aceite no domina: VPN no cruza cero a ningún Brent razonable
            return np.nan, "BRENT_INSENSITIVE"
        return np.nan, "SIN_RAIZ_RENTABLE" if flo > 0 else "SIN_RAIZ_NO_RENTABLE"
    try:
        raiz = float(brentq(funcion, lo, hi, xtol=1e-4))
    except (ValueError, RuntimeError):
        return np.nan, "NO_CONVERGE"
    if raiz <= lo + 0.01:
        return 0.0, "OK_SIEMPRE_RENTABLE"
    return raiz, "OK"


def breakeven_financiero(d: dict) -> tuple[float, str]:
    """Precio NET donde NPV (EV11) = 0.

    Si el campo no tiene aceite en ningún año activo, el VPN no depende del precio
    Brent → devuelve (nan, 'BRENT_INSENSITIVE') para distinguirlo de un fallo numérico.
    """
    tiene_aceite = bool(np.any(d["ACEITE_VTS_NETO"] > 0))
    return _raiz(lambda p: _cadena(p, d)[0], brent_insensible=not tiene_aceite)


def breakeven_operacional(d: dict) -> tuple[float, str]:
    """Precio NET donde el FC del último año con aceite = 0 (límite económico / abandono).

    Usar el último año CON aceite (no el último de la serie) hace el criterio robusto
    ante colas de gas donde fc[-1] sería constante en precio y jamás cruzaría cero.
    Si el campo no tiene aceite en ningún año activo → (nan, 'BRENT_INSENSITIVE').
    """
    if d.get("_ult_oil", -1) < 0:
        return np.nan, "BRENT_INSENSITIVE"
    return _raiz(lambda p: _cadena(p, d)[1])


# ── Perfil de salida volumétrico (directriz §4bis.6, track Calidad) ──────────────

PERFIL_PERCENTILES = (10, 25, 50, 75, 90)


def perfil_salida(d: dict, percentiles=PERFIL_PERCENTILES) -> dict:
    """Precio al que muere el X% del volumen de aceite de la clase (perfil de salida).

    Reemplaza el acantilado de clase (un solo BK del año marginal) por la curva
    acumulada "% del volumen que deja de ser económico" vs precio. Cada AÑO del FC
    tiene su propio precio de quiebre:
        BK_año = (opex + capex − ingreso_gas − ingreso_glp) / aceite_año
    Al bajar el precio, los años mueren del más caro (cola depletada) al más barato.
    BK_Pxx = precio al que el xx% acumulado del volumen ya no es económico.

    Cero datos nuevos: usa el MISMO FC certificado que el breakeven (auditoría s6).
    Retorna {} si el campo no tiene aceite (sin perfil de precio posible).
    """
    ult = d.get("_ult", -1)
    if ult < 0 or d.get("_ult_oil", -1) < 0:
        return {}
    s = slice(0, ult + 1)
    aceite  = d["ACEITE_VTS_NETO"][s]
    opex    = (d["OPEX"][s] + d["OTROS_ING"][s]) * (1 - SENSIBILIDAD_OPEX)
    capex   = d["CAPEX"][s]
    ing_gas = d["GAS_VTS_NETO"][s] * d["PODER_CALOR"][s] * d["PRECIO_GAS"][s]
    ing_glp = d["GLP_NETO"][s] * d["PRECIO_GLP"][s]

    mask = aceite > 0
    if not np.any(mask):
        return {}
    with np.errstate(divide="ignore", invalid="ignore"):
        bk = (opex[mask] + capex[mask] - ing_gas[mask] - ing_glp[mask]) / aceite[mask]
    # Año con ingreso no-aceite ≥ costos → rentable a cualquier precio (nunca sale por
    # precio); su umbral de salida es 0, no negativo.
    bk = np.maximum(bk, 0.0)
    vol = aceite[mask]

    orden   = np.argsort(-bk)              # el año más caro muere primero (precio alto)
    bk_ord  = bk[orden]
    frac    = np.cumsum(vol[orden]) / vol.sum()

    out = {}
    for pct in percentiles:
        j = int(np.searchsorted(frac, pct / 100.0))
        j = min(j, len(bk_ord) - 1)
        out[f"BK_P{pct}"] = float(bk_ord[j])
    return out


# ── Procesamiento de archivos / carpetas ────────────────────────────────────────

def procesar_fc(ruta: Path, retener_datos: bool = False) -> list[dict]:
    """Procesa un FC (un campo) → una fila por clase de reserva presente.

    retener_datos=True: cada fila lleva la clave privada "_DATOS" con los arrays de la
    hoja (o None si SIN_RESERVAS) para que el runner pueda combinar componentes
    multi-FC de un UNIFICADO (combinar_hojas) sin releer los Excel. El runner debe
    descartar "_DATOS" antes de persistir."""
    ruta = Path(ruta)
    crudo, campo = campo_desde_nombre(ruta.name)
    vigencia = vigencia_desde_nombre(ruta.name)
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    filas = []
    for hoja in HOJAS_CLASE:
        if hoja not in wb.sheetnames:
            continue
        ws = wb[hoja]
        offset = detectar_offset(ws)
        cols = resolver_columnas(offset)
        d = leer_hoja(ws, cols)
        if d is None or d["_ult"] < 0:
            fila = _fila(ruta.name, campo, vigencia, hoja, np.nan, np.nan,
                         "SIN_RESERVAS", "SIN_RESERVAS")
            if retener_datos:
                fila["_DATOS"] = None
            filas.append(fila)
            continue
        fin, msg_fin = breakeven_financiero(d)
        ope, msg_ope = breakeven_operacional(d)
        fila = _fila(ruta.name, campo, vigencia, hoja, fin, ope, msg_fin, msg_ope)
        if retener_datos:
            fila["_DATOS"] = d
        filas.append(fila)
    wb.close()
    return filas


MENSAJES_OK = {"OK", "OK_SIEMPRE_RENTABLE"}   # raíz válida (borde inferior = equilibrio 0)


def _fila(archivo, campo, vigencia, clase, fin, ope, msg_fin, msg_ope) -> dict:
    # BRENT_INSENSITIVE: el campo no depende del precio Brent (ingreso de gas fijo domina
    # o no hay aceite). Cuando True, NO se inyectan puntos sintéticos para ese campo.
    es_insensible = (msg_fin == "BRENT_INSENSITIVE" or msg_ope == "BRENT_INSENSITIVE")
    return {
        "ARCHIVO_ORIGEN":                 archivo,
        "CAMPO":                          campo,
        "VIGENCIA_BREAKEVEN":             vigencia,
        "CLASE_RESERVA":                  clase,
        "BREAKEVEN_FINANCIERO_USD_BBL":   fin,
        "BREAKEVEN_OPERACIONAL_USD_BBL":  ope,
        "MENSAJE_FIN":                    msg_fin,
        "MENSAJE_OPE":                    msg_ope,
        "SOLVER_S1_FALLO":                msg_ope not in MENSAJES_OK,   # compat con leer_breakeven
        "BRENT_INSENSITIVE":              es_insensible,
    }


def procesar_carpeta(carpeta) -> list[dict]:
    """Procesa todos los `*.xlsx` de una carpeta (un FC por campo)."""
    filas = []
    for ruta in sorted(Path(carpeta).glob("*.xlsx")):
        filas.extend(procesar_fc(ruta))
    return filas

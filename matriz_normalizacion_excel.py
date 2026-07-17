"""
matriz_normalizacion_excel.py — Generador Excel offline que replica el layout de la Hoja1 del
Consolidado (análisis del usuario) con FÓRMULAS VIVAS (verificable celda a celda, no valores
pegados) + gráficas, una hoja por campo.

Semántica (decisión 2026-07-09, ver experimento_normalizacion.py):
  Baseline   = cierre OFICIAL del año ANTERIOR (A−1)  ← lo que normaliza Hoja1
  Checkpoint = cierre OFICIAL del presente año (A)     ← informativo (muestra el confound)

Diferencia auditada vs Hoja1: Hoja1 ancla el precio/Brent al deck-Q4 del año anterior
(p.ej. 66.666); aquí se usa el cierre HIST EXACTO de A−1 (p.ej. 66.636 neto) como fuente, y las
celdas Baseline lo referencian con fórmulas. El bloque "Cierre HIST A−1" (abajo en cada hoja)
expone esas anclas para trazabilidad.

Campos: gate dorado + materiales con SENSIBILIDAD_NO_IDENTIFICADA y baseline ≥20 MBPE, aplicando
las exclusiones permanentes del piloto (filiales, PAUTO SUR, gas).

Memoria: NO correr con resultados/matriz_normalizacion.xlsx abierto en Excel (lock → crash).

Salida: resultados/matriz_normalizacion.xlsx
"""

import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import Reference, Series, ScatterChart
from openpyxl.chart.marker import Marker
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from homologacion import Homologador

BASE_DIR   = Path(__file__).parent
STAGING    = BASE_DIR / "datos" / "staging"
RESULTADOS = BASE_DIR / "resultados"

FEATURE   = "PRECIO_NETO_USD_BBL"
BRENT     = "BRENT_FLAT_USD_BBL"
VOL       = "VOLUMEN_1P_SENSIBILIDAD_MBPE"
CHECKPOINT = "CHECKPOINT_1P_MBPE"   # cierre A (renombrado en el pipeline 2026-07-09)
OFICIAL   = "VOLUMEN_1P_OFICIAL_MBPE"

# Gate Dorado = pareto-10 (directriz 2026-07-09; ver docs/NORTE.md)
GATE_DORADO = ["RUBIALES", "CASTILLA", "CAÑO SUR ESTE", "CASTILLA NORTE", "AKACIAS",
               "CHICHIMENE", "CHICHIMENE SW", "LA CIRA", "CUPIAGUA", "YARIGUI-CANTAGALLO"]
EXCLUIR_EXPLICITO = {"PAUTO SUR", "BALLENA", "CHUCHUPA"}

VIGENCIAS = ["2024_Q1", "2024_Q2", "2024_Q3", "2024_Q4",
             "2025_Q1", "2025_Q2", "2025_Q3", "2025_Q4", "2026_Q1", "2026_Q2"]
COL0 = 3  # primera columna de datos (C)

# Filas del layout (espejo de Hoja1)
R_VIG, R_P_DECK, R_P_BASE, R_P_VAR = 6, 7, 8, 9
R_R_DECK, R_R_BASE, R_R_VAR, R_CHK = 11, 12, 13, 14
R_B_DECK, R_B_BASE, R_B_VAR = 15, 16, 17
R_ANCLA_TIT, R_ANCLA_YEAR, R_ANCLA_P, R_ANCLA_R, R_ANCLA_B = 20, 21, 22, 23, 24

HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF")
LBL_FONT = Font(bold=True)
PCT_FMT  = "0.0%"
NUM_FMT  = "0.00"


def cargar():
    df = pd.read_parquet(STAGING / "tablon_unico.parquet")
    base = df[df["ES_BASELINE"] & df[OFICIAL].notna()]
    oficial = {(c, int(a)): float(v) for c, a, v in zip(base["CAMPO"], base["AÑO"], base[OFICIAL])}
    pneto   = {(c, int(a)): float(v) for c, a, v in zip(base["CAMPO"], base["AÑO"], base[FEATURE])}
    brent   = {(c, int(a)): float(v) for c, a, v in zip(base["CAMPO"], base["AÑO"], base[BRENT])}
    cons = df[~df["ES_BASELINE"] & ~df["ES_SINTETICO"]
              & df["VIGENCIA"].astype(str).isin(VIGENCIAS)].copy()
    return df, oficial, pneto, brent, cons


def campos_objetivo(df):
    m = pd.read_csv(RESULTADOS / "output_matriz_prediccion.csv").drop_duplicates("CAMPO")
    sni = m[(m["SENSIBILIDAD_NO_IDENTIFICADA"] == True)
            & (m["VOLUMEN_1P_BASELINE_MBPE"] >= 20)]["CAMPO"].tolist()
    h = Homologador()
    insens = (df.groupby("CAMPO")["BRENT_INSENSITIVE"]
                .apply(lambda s: bool(s.dropna().astype(bool).any())))

    def excluido(c):
        return (c in EXCLUIR_EXPLICITO
                or str(h.atributos(c).get("NEW GERENCIA", "")).strip().upper() == "FILIAL"
                or bool(insens.get(c, False)))

    return [c for c in dict.fromkeys(GATE_DORADO + sni) if not excluido(c)]


def _ancla_col_letter(vig):
    """Columna del bloque ANCLA (C=2023, D=2024, E=2025) según A−1 de la vigencia."""
    A = int(vig[:4])
    return get_column_letter(COL0 + (A - 1) - 2023)


def escribir_hoja(ws, campo, oficial, pneto, brent, cons):
    ws["A1"] = campo
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Normalización %Δ contra cierre del año anterior (Baseline A−1) — espejo Hoja1"
    ws["A3"].font = Font(italic=True, size=9)

    etiquetas = {
        R_P_DECK: "Precio Neto (deck)", R_P_BASE: "Precio Baseline (cierre A−1)",
        R_P_VAR: "% var precio", R_R_DECK: "Reservas 1P (deck)",
        R_R_BASE: "Reservas Baseline (cierre A−1)", R_R_VAR: "% var reservas",
        R_CHK: "Checkpoint (cierre A, informativo)", R_B_DECK: "BRENT (deck)",
        R_B_BASE: "BRENT Baseline (cierre A−1)", R_B_VAR: "% var BRENT",
    }
    for r, txt in etiquetas.items():
        ws.cell(r, 1, txt).font = LBL_FONT
    ws.column_dimensions["A"].width = 30

    sub = cons[cons["CAMPO"] == campo].set_index("VIGENCIA")
    for j, vig in enumerate(VIGENCIAS):
        col = COL0 + j
        cl = get_column_letter(col)
        A = int(vig[:4])
        c = ws.cell(R_VIG, col, vig); c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center")
        row = sub.loc[vig] if vig in sub.index else None

        def val(colname):
            if row is None:
                return None
            v = row[colname]
            return float(v) if pd.notna(v) else None

        # Deck (fuente = números)
        for r, colname in [(R_P_DECK, FEATURE), (R_R_DECK, VOL),
                           (R_CHK, CHECKPOINT), (R_B_DECK, BRENT)]:
            v = val(colname)
            cell = ws.cell(r, col, v)
            cell.number_format = NUM_FMT

        # Baseline (fórmula viva → bloque ANCLA)
        acl = _ancla_col_letter(vig)
        ws.cell(R_P_BASE, col, f"=${acl}${R_ANCLA_P}").number_format = NUM_FMT
        ws.cell(R_R_BASE, col, f"=${acl}${R_ANCLA_R}").number_format = NUM_FMT
        ws.cell(R_B_BASE, col, f"=${acl}${R_ANCLA_B}").number_format = NUM_FMT

        # % var (fórmula viva)
        ws.cell(R_P_VAR, col, f"=IF({cl}{R_P_BASE}=0,\"\",({cl}{R_P_DECK}-{cl}{R_P_BASE})/{cl}{R_P_BASE})").number_format = PCT_FMT
        ws.cell(R_R_VAR, col, f"=IF({cl}{R_R_BASE}=0,\"\",({cl}{R_R_DECK}-{cl}{R_R_BASE})/{cl}{R_R_BASE})").number_format = PCT_FMT
        ws.cell(R_B_VAR, col, f"=IF({cl}{R_B_BASE}=0,\"\",({cl}{R_B_DECK}-{cl}{R_B_BASE})/{cl}{R_B_BASE})").number_format = PCT_FMT

    # Bloque ANCLA (cierres HIST por año A−1: 2023, 2024, 2025)
    ws.cell(R_ANCLA_TIT, 1, "Cierre HIST A−1 (fuente HIST 1P.xlsx)").font = LBL_FONT
    ws.cell(R_ANCLA_P, 1, "Precio Neto cierre").font = LBL_FONT
    ws.cell(R_ANCLA_R, 1, "Reservas 1P cierre").font = LBL_FONT
    ws.cell(R_ANCLA_B, 1, "BRENT cierre").font = LBL_FONT
    for k, año in enumerate([2023, 2024, 2025]):
        col = COL0 + k
        ws.cell(R_ANCLA_YEAR, col, str(año)).font = HDR_FONT
        ws.cell(R_ANCLA_YEAR, col).fill = HDR_FILL
        ws.cell(R_ANCLA_P, col, pneto.get((campo, año))).number_format = NUM_FMT
        ws.cell(R_ANCLA_R, col, oficial.get((campo, año))).number_format = NUM_FMT
        ws.cell(R_ANCLA_B, col, brent.get((campo, año))).number_format = NUM_FMT

    _agregar_charts(ws, campo)


def _serie(ws, xrow, yrow, c0, c1, titulo):
    xref = Reference(ws, min_col=c0, max_col=c1, min_row=xrow, max_row=xrow)
    yref = Reference(ws, min_col=c0, max_col=c1, min_row=yrow, max_row=yrow)
    s = Series(yref, xref, title=titulo)
    s.marker = Marker(symbol="circle", size=7)
    s.graphicalProperties.line.noFill = True   # solo marcadores (scatter puro)
    return s


def _scatter(ws, xrow, yrow, x_titulo, titulo):
    ch = ScatterChart()
    ch.title = titulo
    ch.x_axis.title = x_titulo
    ch.y_axis.title = "% var reservas 1P"
    ch.height = 7.5; ch.width = 12
    # Series por vigencia-año (color por año, como Hoja1)
    for (c0, c1, nom) in [(COL0, COL0 + 3, "2024"), (COL0 + 4, COL0 + 7, "2025"),
                          (COL0 + 8, COL0 + 9, "2026")]:
        ch.series.append(_serie(ws, xrow, yrow, c0, c1, nom))
    return ch


def _agregar_charts(ws, campo):
    ws.add_chart(_scatter(ws, R_P_VAR, R_R_VAR, "% var precio neto",
                          "%ΔReservas vs %ΔPrecio Neto"), "N6")
    ws.add_chart(_scatter(ws, R_B_VAR, R_R_VAR, "% var BRENT",
                          "%ΔReservas vs %ΔBRENT"), "N22")
    # Chart (c): todos los quarters juntos (una serie)
    ch = ScatterChart()
    ch.title = "%ΔReservas vs %ΔPrecio (todos los quarters)"
    ch.x_axis.title = "% var precio neto"; ch.y_axis.title = "% var reservas 1P"
    ch.height = 7.5; ch.width = 12
    ch.series.append(_serie(ws, R_P_VAR, R_R_VAR, COL0, COL0 + 9, "Todos"))
    ws.add_chart(ch, "N38")


def hoja_resumen(wb, campos, oficial, pneto, cons):
    ws = wb.create_sheet("RESUMEN", 0)
    cols = ["CAMPO", "N_Q", "CORR_PRECIO_RES", "ELAST_SLOPE", "SOLAPA_PCT",
            "Q_MIN_PPRECIO", "PPRECIO_MIN", "PRES_EN_MIN"]
    for j, h in enumerate(cols, 1):
        c = ws.cell(1, j, h); c.font = HDR_FONT; c.fill = HDR_FILL
    ws.column_dimensions["A"].width = 22

    for i, campo in enumerate(campos, 2):
        sub = cons[cons["CAMPO"] == campo].set_index("VIGENCIA")
        pp, pr, años, vigs = [], [], [], []
        for vig in VIGENCIAS:
            if vig not in sub.index:
                continue
            A = int(vig[:4])
            base_p = pneto.get((campo, A - 1)); base_r = oficial.get((campo, A - 1))
            r = sub.loc[vig]
            if (base_p is None or base_r is None or pd.isna(base_p) or pd.isna(base_r)
                    or not base_p or not base_r or pd.isna(r[FEATURE]) or pd.isna(r[VOL])):
                continue
            pp.append((r[FEATURE] - base_p) / base_p)
            pr.append((r[VOL] - base_r) / base_r)
            años.append(A); vigs.append(vig)
        ws.cell(i, 1, campo)
        n = len(pp)
        ws.cell(i, 2, n)
        if n >= 2:
            pp_a, pr_a, an = np.array(pp), np.array(pr), np.array(años)
            corr = float(np.corrcoef(pp_a, pr_a)[0, 1]) if pp_a.std() > 1e-9 else np.nan
            slope = float(np.polyfit(pp_a, pr_a, 1)[0]) if pp_a.std() > 1e-9 else np.nan
            # solape de bandas %Δprecio entre años
            us = np.unique(an); mejor = np.nan
            if len(us) >= 2:
                bandas = {a: (pp_a[an == a].min(), pp_a[an == a].max()) for a in us}
                mejor = 0.0
                for k, a in enumerate(us):
                    for b in us[k + 1:]:
                        lo = max(bandas[a][0], bandas[b][0]); hi = min(bandas[a][1], bandas[b][1])
                        mejor = max(mejor, hi - lo)
            jmin = int(np.argmin(pp_a))
            ws.cell(i, 3, round(corr, 3) if pd.notna(corr) else None)
            ws.cell(i, 4, round(slope, 3) if pd.notna(slope) else None)
            ws.cell(i, 5, round(float(mejor), 4) if pd.notna(mejor) else None)
            ws.cell(i, 6, vigs[jmin])
            ws.cell(i, 7, round(float(pp_a[jmin]), 4))
            ws.cell(i, 8, round(float(pr_a[jmin]), 4))


def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    df, oficial, pneto, brent, cons = cargar()
    campos = campos_objetivo(df)
    print(f"Campos ({len(campos)}): {campos}")

    wb = Workbook()
    wb.remove(wb.active)
    for campo in campos:
        nombre = campo[:31].replace("/", "-").replace("\\", "-")
        ws = wb.create_sheet(nombre)
        escribir_hoja(ws, campo, oficial, pneto, brent, cons)
    hoja_resumen(wb, campos, oficial, pneto, cons)

    salida = RESULTADOS / "matriz_normalizacion.xlsx"
    wb.save(salida)
    print(f"Guardado: {salida}  ({len(campos)} hojas + RESUMEN)")


if __name__ == "__main__":
    main()
